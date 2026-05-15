#!/usr/bin/env python3
"""E-learning PDF to Thinkific XLSX quiz converter.

Extracts text from e-learning PDF modules, generates quiz questions
(via AI or manually), and exports to Thinkific import format (.xlsx).

Backend module used by both the CLI and the Streamlit GUI (app.py).
"""

import json
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font

NOISE_PATTERNS = [
    r"^afbeelding:.*$",
    r"^Welkom\s*$",
    r"^Samenvatting\s*$",
    r"^Over deze module\s*$",
]

SYSTEM_PROMPT = """Je bent een expert in het maken van quizvragen op basis van Nederlandse e-learning materialen. Je werkt onder strenge anti-plagiaat richtlijnen.

Je krijgt de volledige tekst van een e-learning PDF module. Je taak: maak quizvragen die de **kennis en leerdoelen** uit de tekst toetsen, in volledig eigen formulering.

== STRIKT VERBOD OP PLAGIAAT ==

De bron-PDF is auteursrechtelijk beschermd. Jouw output wordt door plagiaatdetectie-software gecontroleerd. Eerdere generaties scoorden te hoog op plagiaat. Dit MOET nu beter. Volg deze regels rigoureus:

**1. WOORDKEUZE — gebruik synoniemen en parafraseer**
- Vervang werkwoorden waar mogelijk: "uitvoeren" → "verrichten/toepassen/inzetten", "controleren" → "nagaan/verifiëren", "zorgen voor" → "regelen/waarborgen", enz.
- Vervang voegwoorden en zinsconnectoren: vermijd dat zinnen letterlijk zo beginnen als in de bron
- Vakbegrippen, eigennamen, exacte normen (cijfers, meeteenheden, wettelijke termen) MOGEN identiek blijven — die zijn feitelijk

**2. ZINSOPBOUW — kies een andere grammaticale structuur**
- Verwissel hoofd- en bijzinnen
- Maak van een bevestigende zin een vraagzin, of omgekeerd
- Splits lange bronzinnen op in twee korte vragen, of combineer juist twee feiten in één vraag
- Begin nooit met dezelfde woorden als de bronzin

**3. CASUS / VOORBEELDEN — verzin eigen, nieuwe scenario's**
- Als de PDF een voorbeeld geeft (bv. "een auto rijdt het water in bij de pylon om 18.15"), gebruik dat voorbeeld NIET. Verzin een ander scenario dat dezelfde kennis toetst.
- Wissel locatie, tijdstip, betrokken personen, omstandigheden
- Een vraag moet hetzelfde leerdoel toetsen, niet dezelfde situatie navertellen

**4. HARDE LIMIETEN**
- Maximaal 3 opeenvolgende woorden uit de bron mogen identiek voorkomen in je output (uitgezonderd vakbegrippen en eigennamen)
- Geen enkele zin in je vraag, uitleg of antwoord mag een directe knip-en-plakzin uit de bron zijn
- Als je merkt dat een formulering te dicht bij de bron komt: herschrijf hem volledig

**5. ZELF-CONTROLE VOOR ELKE VRAAG**
Voordat je een vraag opneemt, controleer mentaal:
- Zou een plagiaatdetector dit als overgenomen markeren? Zo ja: herschrijf.
- Lijkt mijn zinsstructuur op een zin uit de bron? Zo ja: kies een andere structuur.
- Gebruik ik dezelfde casus/voorbeeld als de bron? Zo ja: verzin een nieuwe.

== UITGEWERKT VOORBEELD ==

**Bron-PDF zegt:**
"Een grijpredding mag uitgevoerd worden als het slachtoffer maximaal 15 meter van de wal of boot ligt en het water maximaal 1,5 meter diep is."

❌ **FOUT (te letterlijk, plagiaat-risico):**
  Vraag: "Wanneer mag een grijpredding uitgevoerd worden?"
  Antwoord: "Als het slachtoffer maximaal 15 meter van de wal of boot ligt en het water maximaal 1,5 meter diep is."
  → Reden: complete bronzin overgenomen.

❌ **FOUT (lichte variatie, nog steeds plagiaat):**
  Vraag: "Een grijpredding mag uitgevoerd worden tot welke afstand en diepte?"
  Antwoord: "Maximaal 15 meter van de wal of boot en water maximaal 1,5 meter diep."
  → Reden: zinsstructuur en woordvolgorde te dicht bij origineel.

✅ **GOED (eigen scenario, andere structuur, parafrase):**
  Vraag: "Tijdens een oefening tref je een drenkeling 12 meter uit de kade aan, in water dat ongeveer een meter diep is. Is een grijpredding hier een toegestane optie voor een individuele manschap?"
  Antwoorden: "Ja, deze afstand en diepte vallen binnen de grenzen", "Nee, de afstand is te groot", "Nee, alleen een oppervlakteredding mag hier", "Nee, alleen een duikinzet is toegestaan"
  Correct: het eerste antwoord
  Uitleg: "De norm laat een individuele grijpredding toe binnen 15 m afstand en 1,5 m diepte; deze situatie blijft daar ruim onder."
  → Eigen casus (oefening + kade), andere zinsstructuur (scenario + ja/nee-vraag), eigen woordkeuze ("kade" ipv "wal", "norm laat toe" ipv "mag uitgevoerd worden").

== WAT JE MOET DOEN ==

1. Identificeer de belangrijkste leerdoelen en feiten in de tekst (definities, normen, procedures, oorzaken/gevolgen, beslismomenten, risico's).

2. Maak voor elk leerdoel een quizvraag in volledig eigen formulering. Wissel af tussen:
   - Praktijkscenario's (met door jou verzonnen context)
   - Toepassingsvragen ("Welke aanpak past hier?")
   - Vergelijkingsvragen ("Wat onderscheidt X van Y?")
   - Begripsvragen (parafraseer de definitie)
   - Beslissingsvragen ("Wat doe je als...?")

3. Voor elke vraag, lever een JSON-object:
   - "type": "SA" (één correct antwoord) of "MA" (meerdere correcte antwoorden)
   - "question": de vraagtekst, eigen formulering
   - "explanation": korte uitleg (1-2 zinnen), eigen formulering
   - "choices": lijst van {"text": "...", "correct": true/false}
     - Minimaal 3, maximaal 6 antwoorden per vraag
     - Bij SA: precies 1 correct, bij MA: 2 of meer correct
     - Afleidende antwoorden moeten plausibel zijn, niet flauw
     - Antwoorden eveneens in eigen formulering

4. Baseer correcte antwoorden UITSLUITEND op informatie die werkelijk in de lesstof staat. Verzin geen feiten of normen.

5. Negeer navigatie-elementen, headers, footers en afbeeldingsbeschrijvingen.

Antwoord ALLEEN met een JSON-array van vraag-objecten. Geen andere tekst, geen markdown code blocks."""

USER_PROMPT_TEMPLATE = """Maak quizvragen op basis van onderstaande e-learning lesstof.

KRITIEKE INSTRUCTIE: De output wordt door plagiaatsoftware gecontroleerd. Schrijf elke vraag, elk antwoord en elke uitleg in **volledig eigen formulering**:
- Eigen woordkeuze (synoniemen, parafrase)
- Eigen zinsstructuur (andere grammaticale opbouw dan de bron)
- Eigen casussen/voorbeelden (NIET de scenario's uit de bron hergebruiken)
- Maximaal 3 opeenvolgende woorden identiek aan de bron

Cijfers, normen, vakbegrippen en eigennamen mogen wel identiek blijven (die zijn feitelijk).

=== BEGIN LESSTOF ===
{text}
=== EINDE LESSTOF ===

Antwoord met een JSON-array van vraag-objecten:"""

AI_PROVIDERS = {
    "claude": {
        "name": "Claude (Anthropic)",
        "default_model": "claude-sonnet-4-6",
        "models": ["claude-sonnet-4-6", "claude-haiku-4-5-20251001", "claude-opus-4-6"],
        "env_key": "ANTHROPIC_API_KEY",
        "key_url": "https://console.anthropic.com/settings/keys",
        "needs_api_key": True,
    },
    "openai": {
        "name": "OpenAI (ChatGPT)",
        "default_model": "gpt-4o",
        "models": ["gpt-4o", "gpt-4o-mini", "gpt-4.1"],
        "env_key": "OPENAI_API_KEY",
        "key_url": "https://platform.openai.com/api-keys",
        "needs_api_key": True,
    },
    "gemini": {
        "name": "Google Gemini",
        "default_model": "gemini-2.5-flash",
        "models": ["gemini-2.5-flash", "gemini-2.5-pro", "gemini-2.0-flash"],
        "env_key": "GEMINI_API_KEY",
        "key_url": "https://aistudio.google.com/apikey",
        "needs_api_key": True,
    },
    "ollama": {
        "name": "Ollama (lokaal)",
        "default_model": "llama3.1",
        "models": ["llama3.1", "llama3.2", "mistral", "gemma2"],
        "env_key": None,
        "key_url": None,
        "needs_api_key": False,
    },
}


# === PDF text extraction ===

def extract_text(pdf_path: str) -> str:
    pages = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")
            filtered = []
            for line in lines:
                skip = False
                for pattern in NOISE_PATTERNS:
                    if re.match(pattern, line.strip()):
                        skip = True
                        break
                if not skip and line.strip():
                    filtered.append(line.strip())
            if filtered:
                pages.append("\n".join(filtered))
    return "\n\n---\n\n".join(pages)


def extract_text_from_bytes(pdf_bytes: bytes) -> str:
    import io
    pages = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            lines = text.split("\n")
            filtered = []
            for line in lines:
                skip = False
                for pattern in NOISE_PATTERNS:
                    if re.match(pattern, line.strip()):
                        skip = True
                        break
                if not skip and line.strip():
                    filtered.append(line.strip())
            if filtered:
                pages.append("\n".join(filtered))
    return "\n\n---\n\n".join(pages)


# === AI question generation ===

def _parse_json_response(content: str) -> list[dict]:
    match = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", content, re.DOTALL)
    if match:
        content = match.group(1)
    return json.loads(content)


def call_claude(text: str, model: str, api_key: str = None) -> list[dict]:
    import anthropic
    kwargs = {"api_key": api_key} if api_key else {}
    client = anthropic.Anthropic(**kwargs)
    response = client.messages.create(
        model=model,
        max_tokens=8192,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": USER_PROMPT_TEMPLATE.format(text=text)}],
    )
    return _parse_json_response(response.content[0].text)


def call_openai(text: str, model: str, api_key: str = None) -> list[dict]:
    import openai
    kwargs = {"api_key": api_key} if api_key else {}
    client = openai.OpenAI(**kwargs)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": USER_PROMPT_TEMPLATE.format(text=text)},
        ],
        temperature=0.3,
    )
    return _parse_json_response(response.choices[0].message.content)


def call_gemini(text: str, model: str, api_key: str = None) -> list[dict]:
    from google import genai
    kwargs = {"api_key": api_key} if api_key else {}
    client = genai.Client(**kwargs)
    response = client.models.generate_content(
        model=model,
        contents=SYSTEM_PROMPT + "\n\n" + USER_PROMPT_TEMPLATE.format(text=text),
    )
    return _parse_json_response(response.text)


def call_ollama(text: str, model: str, api_key: str = None) -> list[dict]:
    import urllib.request
    payload = json.dumps({
        "model": model,
        "stream": False,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": USER_PROMPT_TEMPLATE.format(text=text)},
        ],
    }).encode()
    req = urllib.request.Request(
        "http://localhost:11434/api/chat",
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    with urllib.request.urlopen(req, timeout=300) as resp:
        result = json.loads(resp.read())
    return _parse_json_response(result["message"]["content"])


def generate_questions(text: str, provider: str, model: str = None, api_key: str = None) -> list[dict]:
    info = AI_PROVIDERS[provider]
    model = model or info["default_model"]
    fn = {"claude": call_claude, "openai": call_openai, "gemini": call_gemini, "ollama": call_ollama}[provider]
    return fn(text, model, api_key)


# === JSON import ===

def parse_questions_json(content: str) -> list[dict]:
    match = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", content, re.DOTALL)
    if match:
        content = match.group(1)
    return json.loads(content)


# === Prompt builder ===

def build_prompt(text: str) -> str:
    return SYSTEM_PROMPT + "\n\n---\n\n" + USER_PROMPT_TEMPLATE.format(text=text)


# === XLSX export ===

def export_xlsx(questions: list[dict], output_path: str):
    wb = Workbook()

    ws = wb.active
    ws.title = "QUESTIONS"
    headers = ["QuestionType", "QuestionText", "Explanation"] + [
        f"Choice{i}" for i in range(1, 11)
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for q in questions:
        row = [q.get("type", "SA"), q["question"], q.get("explanation", "")]
        choices = q.get("choices", [])
        for choice in choices[:10]:
            text = choice["text"]
            if choice.get("correct"):
                text = f"*{text}"
            row.append(text)
        row.extend([""] * (13 - len(row)))
        ws.append(row)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 60
    for col_letter in "DEFGHIJKLM":
        ws.column_dimensions[col_letter].width = 40

    ws2 = wb.create_sheet("INSTRUCTIONS")
    instructions = [
        ["Question Type", "Question Type is required."],
        ["", "Use the following abbreviations for supported question types:"],
        ["", "SA", "One correct answer (radio button)"],
        ["", "MA", "One or more correct answers (checkboxes)"],
        ["Question Text", "Question Text is required for all questions."],
        [],
        ["Answer Choices", "Answer Choices are required for all question types"],
        ["", "Enter up to 10 answer choices for each question."],
        ["", "Designate correct answers with an asterisk (*) at the beginning.\nExample: *True"],
        ["Notes", "Formatting in your spreadsheet will be removed during import."],
        ["", "Only questions in the first worksheet (tab) will be imported."],
        ["", "Please do not rename the worksheets (tabs) in this file."],
        ["", "Questions will be imported at the end of the quiz."],
    ]
    for row in instructions:
        ws2.append(row)

    wb.save(output_path)


def export_xlsx_to_bytes(questions: list[dict]) -> bytes:
    import io
    wb = Workbook()

    ws = wb.active
    ws.title = "QUESTIONS"
    headers = ["QuestionType", "QuestionText", "Explanation"] + [
        f"Choice{i}" for i in range(1, 11)
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for q in questions:
        row = [q.get("type", "SA"), q["question"], q.get("explanation", "")]
        choices = q.get("choices", [])
        for choice in choices[:10]:
            text = choice["text"]
            if choice.get("correct"):
                text = f"*{text}"
            row.append(text)
        row.extend([""] * (13 - len(row)))
        ws.append(row)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 60
    for col_letter in "DEFGHIJKLM":
        ws.column_dimensions[col_letter].width = 40

    ws2 = wb.create_sheet("INSTRUCTIONS")
    instructions = [
        ["Question Type", "Question Type is required."],
        ["", "Use the following abbreviations for supported question types:"],
        ["", "SA", "One correct answer (radio button)"],
        ["", "MA", "One or more correct answers (checkboxes)"],
        ["Question Text", "Question Text is required for all questions."],
        [],
        ["Answer Choices", "Answer Choices are required for all question types"],
        ["", "Enter up to 10 answer choices for each question."],
        ["", "Designate correct answers with an asterisk (*) at the beginning.\nExample: *True"],
        ["Notes", "Formatting in your spreadsheet will be removed during import."],
        ["", "Only questions in the first worksheet (tab) will be imported."],
        ["", "Please do not rename the worksheets (tabs) in this file."],
        ["", "Questions will be imported at the end of the quiz."],
    ]
    for row in instructions:
        ws2.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# === CLI (kept for power users) ===

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="E-learning PDF naar Thinkific XLSX quiz converter.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Voorbeelden:
  %(prog)s "Module.pdf" --text-only              Alleen tekst extraheren
  %(prog)s "Module.pdf" --prompt-only             Tekst + AI prompt genereren
  %(prog)s "Module.pdf" --ai claude               Vragen genereren met Claude
  %(prog)s "Module.pdf" --ai openai               Vragen genereren met ChatGPT
  %(prog)s "Module.pdf" --ai gemini               Vragen genereren met Gemini
  %(prog)s "Module.pdf" --ai ollama               Vragen genereren met Ollama
  %(prog)s --from-json vragen.json                JSON importeren naar XLSX

Start de grafische interface:
  streamlit run app.py
        """,
    )

    parser.add_argument("input", nargs="?", help="Input PDF bestand")
    parser.add_argument("--from-json", metavar="FILE", help="Importeer vragen uit JSON bestand")
    parser.add_argument("-o", "--output", help="Output bestand")
    parser.add_argument("--text-only", action="store_true", help="Alleen tekst extraheren")
    parser.add_argument("--prompt-only", action="store_true", help="Tekst + AI prompt genereren")
    parser.add_argument("--ai", choices=list(AI_PROVIDERS.keys()), metavar="BACKEND", help="AI backend")
    parser.add_argument("--model", help="Model naam")
    parser.add_argument("--api-key", help="API key (anders via environment variable)")
    parser.add_argument("--review", action="store_true", default=True, dest="review")
    parser.add_argument("--no-review", action="store_false", dest="review")
    args = parser.parse_args()

    if not args.input and not args.from_json:
        parser.error("Geef een PDF bestand op, of gebruik --from-json")

    if args.from_json:
        json_path = Path(args.from_json)
        if not json_path.exists():
            print(f"Fout: bestand niet gevonden: {json_path}")
            sys.exit(1)
        with open(json_path, "r", encoding="utf-8") as f:
            questions = parse_questions_json(f.read())
        print(f"{len(questions)} vragen geladen.")
        if args.review:
            questions = _cli_review(questions)
        output_path = args.output or str(json_path.with_suffix("")) + "_thinkific.xlsx"
        export_xlsx(questions, output_path)
        print(f"Export voltooid: {output_path} ({len(questions)} vragen)")
        return

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Fout: bestand niet gevonden: {input_path}")
        sys.exit(1)

    print(f"PDF tekst extraheren uit: {input_path}")
    text = extract_text(str(input_path))
    print(f"Tekst geextraheerd: {len(text)} tekens")

    if args.text_only:
        output_path = args.output or str(input_path.with_suffix("")) + "_tekst.txt"
        Path(output_path).write_text(text, encoding="utf-8")
        print(f"Tekst opgeslagen: {output_path}")
        return

    if args.prompt_only:
        prompt = build_prompt(text)
        output_path = args.output or str(input_path.with_suffix("")) + "_prompt.txt"
        Path(output_path).write_text(prompt, encoding="utf-8")
        print(f"AI prompt opgeslagen: {output_path}")
        print(f"\nKopieer naar Claude.ai, ChatGPT, etc. en gebruik dan:")
        print(f"  python {sys.argv[0]} --from-json <output.json>")
        return

    if not args.ai:
        print("\nKies een modus:")
        print("  --text-only     Alleen tekst exporteren")
        print("  --prompt-only   Tekst + AI prompt genereren")
        for key, info in AI_PROVIDERS.items():
            print(f"  --ai {key:10s} Vragen genereren met {info['name']}")
        print("  --from-json     Vragen importeren uit JSON")
        sys.exit(1)

    print(f"Vragen genereren met {AI_PROVIDERS[args.ai]['name']}...")
    try:
        questions = generate_questions(text, args.ai, args.model, args.api_key)
    except Exception as e:
        print(f"\nFout: {e}")
        info = AI_PROVIDERS[args.ai]
        if info["needs_api_key"]:
            print(f"Stel {info['env_key']} in, of gebruik --api-key")
            print(f"API key aanmaken: {info['key_url']}")
        elif args.ai == "ollama":
            print("Zorg dat Ollama draait: ollama serve")
        sys.exit(1)

    print(f"{len(questions)} vragen gegenereerd.")
    if args.review:
        questions = _cli_review(questions)
    output_path = args.output or str(input_path.with_suffix("")) + "_thinkific.xlsx"
    export_xlsx(questions, output_path)
    print(f"Export voltooid: {output_path} ({len(questions)} vragen)")


def _cli_review(questions: list[dict]) -> list[dict]:
    print(f"\n{len(questions)} vragen. Review starten...\n")
    accepted = []
    for i, q in enumerate(questions):
        qtype = q.get("type", "SA")
        print(f"\n{'='*60}")
        print(f"  Vraag {i+1}/{len(questions)} [{qtype}]")
        print(f"{'='*60}")
        print(f"  {q['question']}")
        print(f"  Uitleg: {q.get('explanation', '-')}")
        for j, c in enumerate(q.get("choices", []), 1):
            marker = "*" if c.get("correct") else " "
            print(f"  {marker} {j}. {c['text']}")
        while True:
            choice = input("\n  [A]ccepteren / [S]kippen / [E]diten? ").strip().lower()
            if choice in ("a", ""):
                accepted.append(q)
                break
            elif choice == "s":
                break
            elif choice == "e":
                new_text = input("  Nieuwe vraagtekst (leeg=behouden): ").strip()
                if new_text:
                    q["question"] = new_text
                new_expl = input("  Nieuwe uitleg (leeg=behouden): ").strip()
                if new_expl:
                    q["explanation"] = new_expl
                accepted.append(q)
                break
    print(f"\n{len(accepted)} van {len(questions)} vragen geaccepteerd.")
    return accepted


if __name__ == "__main__":
    main()
